#!/usr/bin/env python3
import argparse
import csv
import html
import random
import re
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Tuple
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

import requests
from openpyxl import Workbook, load_workbook

from fruugo_link_tracker import LinkTracker


DEFAULT_TIMEOUT = 30
DATA_START_ROW = 4
IMAGE_COLUMNS = ["Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH"]
EMPTY_COLUMNS = {
    "A", "B", "E", "I", "J", "R", "S", "U", "W", "AN", "AO", "AU", "AV", "AW",
    "AX", "AY", "AZ", "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BJ", "BL", "BN",
    "BO", "BP", "BR", "BS", "BT", "BU", "BV", "BW", "BX"
}
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
)
ERP_PRODUCT_PREFIX = "Prod_1772601378_NEW_FRU_GBR_01_1772601383"
FRUUGO_BASE_URL = "https://www.fruugo.co.uk"
FRUUGO_SITEMAP_URL = f"{FRUUGO_BASE_URL}/sitemap"
PLAYWRIGHT_FETCH_SCRIPT = Path(__file__).with_name("fruugo_fetch_html.js")
DEFAULT_HOT_CATEGORY_URLS = {
    "Humidifiers": f"{FRUUGO_BASE_URL}/humidifiers/a-613",
    "Powered Hand Fans & Misters": f"{FRUUGO_BASE_URL}/powered-hand-fans-misters/a-7527",
    "Jump Ropes": f"{FRUUGO_BASE_URL}/jump-ropes/a-2614",
    "Can Openers": f"{FRUUGO_BASE_URL}/can-openers/a-733",
    "Motor Vehicle Lighting": f"{FRUUGO_BASE_URL}/motor-vehicle-lighting/a-3318",
    "Microphones": f"{FRUUGO_BASE_URL}/microphones/a-234",
    "Home Decor Decals": f"{FRUUGO_BASE_URL}/home-decor-decals/a-3221",
    "Dog Toys": f"{FRUUGO_BASE_URL}/dog-toys/a-5010",
    "Lighting Timers": f"{FRUUGO_BASE_URL}/lighting-timers/a-3522",
    "GPS Tracking Devices": f"{FRUUGO_BASE_URL}/gps-tracking-devices/a-6544",
}
DEFAULT_HOT_CATEGORY_NAMES = list(DEFAULT_HOT_CATEGORY_URLS.keys())
CATEGORY_KEYWORD_HINTS = {
    "Humidifiers": ["humidifier", "humidifiers", "humidair"],
    "Powered Hand Fans & Misters": ["fan", "fans", "mister", "misting", "cooling", "neck fan", "handheld"],
    "Jump Ropes": ["jump rope", "jump", "rope", "skipping"],
    "Can Openers": ["can opener", "opener", "jar opener", "kitchen"],
    "Motor Vehicle Lighting": ["light", "lights", "lighting", "led", "headlight", "tail", "lamp", "driving"],
    "Microphones": ["microphone", "microphones", "mic", "recording", "podcast"],
    "Home Decor Decals": ["decal", "decals", "sticker", "stickers", "wall", "home decor"],
    "Dog Toys": ["dog", "puppy", "pet toy", "treat ball", "chew", "squeaky"],
    "Lighting Timers": ["timer", "countdown", "switch", "lighting"],
    "GPS Tracking Devices": ["gps", "tracker", "tracking", "locator", "vehicle tracker"],
}


@dataclass
class ProductInput:
    group_id: str
    url: str
    is_primary: str = ""
    color: str = ""
    size: str = ""
    attrs: str = ""
    barcode: str = ""
    rrp: str = ""
    category: str = ""
    title: str = ""
    description: str = ""
    images: str = ""


@dataclass
class ProductData:
    title: str
    description: str
    rrp: Decimal
    category: str
    color: str
    size: str
    attrs: str
    barcode: str
    images: List[str]
    source_url: str


@dataclass
class GroupedProduct:
    group_id: str
    rows: List[ProductData]


@dataclass
class CategoryTarget:
    name: str
    url: str


@dataclass
class ProductCandidate:
    category_name: str
    url: str
    rank: int
    title: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Discover hot Fruugo products and generate a Fruugo bulk upload Excel."
    )
    parser.add_argument("--template", help="Path to the Fruugo Excel template.")
    parser.add_argument("--input", help="Path to the source CSV.")
    parser.add_argument("--output", help="Path to the generated Excel file.")
    parser.add_argument("--operator", default="SHOU", help="Operator prefix used in vendor SKU.")
    parser.add_argument("--shop", default="07", help="Shop code used in vendor SKU.")
    parser.add_argument(
        "--date-code",
        default=datetime.now().strftime("%m%d"),
        help="MMDD code used in vendor SKU, default is today."
    )
    parser.add_argument("--vendor-code", default="SAYAV", help="Value for column F.")
    parser.add_argument(
        "--manufacturer",
        default="san ya yan wu xu shang mao you xian gong si",
        help="Value for column BH."
    )
    parser.add_argument("--brand", default="Unbranded", help="Value for column BI.")
    parser.add_argument(
        "--clear-existing",
        action="store_true",
        default=True,
        help="Clear all rows from row 4 downward before writing new rows."
    )
    parser.add_argument(
        "--no-clear-existing",
        action="store_false",
        dest="clear_existing",
        help="Do not clear existing data rows."
    )
    parser.add_argument(
        "--discount-rate",
        default="0.4",
        help="Column BM is RRP multiplied by this rate. Default 0.4."
    )
    parser.add_argument(
        "--inventory-dir",
        default="",
        help="Directory to write the inventory workbook. Defaults to the product output directory."
    )
    parser.add_argument(
        "--inventory-prefix",
        default="SAYA_Inventory_upload",
        help="Filename prefix for the inventory workbook."
    )
    parser.add_argument(
        "--inventory-warehouse",
        default="SAYAW",
        help="Warehouse value for the inventory workbook."
    )
    parser.add_argument(
        "--inventory-code-field-name",
        default="UPC",
        help="codeFieldName value for the inventory workbook."
    )
    parser.add_argument(
        "--inventory-min-qty",
        type=int,
        default=200,
        help="Minimum random quantity for inventory rows."
    )
    parser.add_argument(
        "--inventory-max-qty",
        type=int,
        default=300,
        help="Maximum random quantity for inventory rows."
    )
    parser.add_argument(
        "--discover-hot-count",
        type=int,
        default=0,
        help="Automatically discover this many hot Fruugo products when --input is omitted."
    )
    parser.add_argument(
        "--discover-products-per-category",
        type=int,
        default=3,
        help="How many product links to collect from each category page during discovery."
    )
    parser.add_argument(
        "--discover-categories",
        default=",".join(DEFAULT_HOT_CATEGORY_NAMES),
        help="Comma-separated Fruugo category names used for hot product discovery."
    )
    parser.add_argument(
        "--discover-output-csv",
        default="",
        help="Optional path for saving the discovered hot products CSV."
    )
    parser.add_argument(
        "--discover-only",
        action="store_true",
        help="Only discover hot products and write the CSV, without generating XLSX files."
    )
    parser.add_argument(
        "--tracker-db",
        default="",
        help="Optional SQLite tracker DB used to skip consumed URLs and mark scrape state."
    )
    parser.add_argument(
        "--tracker-source-label",
        default="",
        help="Optional source label written into the tracker when URLs are seen."
    )
    parser.add_argument(
        "--tracker-skip-failed",
        action="store_true",
        help="If set, URLs previously marked as failed will not be retried."
    )
    return parser.parse_args()


def read_input_csv(path: Path) -> List[ProductInput]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"url"}
        fieldnames = {name.strip().lower() for name in (reader.fieldnames or [])}
        missing = required - fieldnames
        if missing:
            raise ValueError(f"Input CSV missing required columns: {', '.join(sorted(missing))}")

        products: List[ProductInput] = []
        for index, row in enumerate(reader, start=2):
            normalized = {str(k).strip().lower(): (v or "").strip() for k, v in row.items()}
            url = normalized.get("url", "")
            if not url:
                continue
            products.append(
                ProductInput(
                    group_id=normalized.get("group_id", "") or normalized.get("url", ""),
                    url=url,
                    is_primary=normalized.get("is_primary", ""),
                    color=normalized.get("color", ""),
                    size=normalized.get("size", ""),
                    attrs=normalized.get("attrs", ""),
                    barcode=normalized.get("barcode", ""),
                    rrp=normalized.get("rrp", ""),
                    category=normalized.get("category", ""),
                    title=normalized.get("title", ""),
                    description=normalized.get("description", ""),
                    images=normalized.get("images", "")
                )
            )
        if not products:
            raise ValueError("Input CSV has no usable product rows.")
        return products


def fetch_html(url: str) -> str:
    try:
        response = requests.get(
            url,
            headers={"User-Agent": USER_AGENT, "Accept-Language": "en-GB,en;q=0.9"},
            timeout=DEFAULT_TIMEOUT
        )
        response.raise_for_status()
        return response.text
    except requests.HTTPError as exc:
        if exc.response is None or exc.response.status_code not in {403, 429}:
            raise
        return fetch_html_with_browser(url)


def fetch_html_with_browser(url: str) -> str:
    if not PLAYWRIGHT_FETCH_SCRIPT.exists():
        raise FileNotFoundError(f"Playwright fetch script not found: {PLAYWRIGHT_FETCH_SCRIPT}")

    completed = subprocess.run(
        ["node", str(PLAYWRIGHT_FETCH_SCRIPT), url],
        capture_output=True,
        check=True,
        text=True,
        timeout=120,
    )
    return completed.stdout


def extract_first(text: str, patterns: List[str], flags: int = 0) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return match.group(1).strip()
    return ""


def clean_text(value: str) -> str:
    value = value.replace("\\r", " ").replace("\\n", " ")
    value = value.replace("&amp;amp;", "&amp;")
    value = html.unescape(value)
    value = re.sub(r"<br\s*/?>", " ", value, flags=re.I)
    value = re.sub(r"<[^>]+>", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_category(category: str) -> str:
    parts = [part.strip() for part in category.split(">") if part.strip()]
    return " > ".join(parts)


def extract_breadcrumb_category(page: str) -> str:
    nav_match = re.search(
        r'<nav[^>]+aria-label=["\']breadcrumbs["\'][^>]*>(.*?)</nav>',
        page,
        re.S | re.I,
    )
    if not nav_match:
        return ""

    breadcrumb_names = []
    for match in re.finditer(
        r'<a[^>]+class=["\'][^"\']*breadcrumb__link[^"\']*["\'][^>]*>(.*?)</a>',
        nav_match.group(1),
        re.S | re.I,
    ):
        name = clean_text(match.group(1))
        if name and name != "Fruugo":
            breadcrumb_names.append(name)

    return normalize_category(" > ".join(breadcrumb_names))


def title_case_words(value: str) -> str:
    if not value:
        return ""
    return " ".join(word[:1].upper() + word[1:] if word else "" for word in value.split())


def random_barcode() -> str:
    return "".join(random.choice("0123456789") for _ in range(12))


def parse_decimal(value: str, fallback: str = "0") -> Decimal:
    raw = value.strip() if value else fallback
    return Decimal(raw)


def derive_attrs(color: str, size: str, attrs: str) -> str:
    if attrs:
        return attrs.upper()
    derived = []
    if color:
        derived.append("COLOR")
    if size:
        derived.append("SIZE")
    return ",".join(derived)


def parse_bool(value: str) -> bool:
    return value.strip().lower() in {"1", "true", "yes", "y", "primary", "main"} if value else False


def source_has_required_fields(source: ProductInput) -> bool:
    return all(
        [
            source.title.strip(),
            source.description.strip(),
            source.category.strip(),
            source.rrp.strip(),
            source.images.strip(),
        ]
    )


def validate_output_filename(path: Path) -> None:
    if not path.name.startswith(ERP_PRODUCT_PREFIX):
        raise ValueError(
            f"Output filename must start with '{ERP_PRODUCT_PREFIX}' to satisfy ERP import rules: {path.name}"
        )
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Output filename must end with .xlsx: {path.name}")

    stem = path.stem
    suffix = stem[len(ERP_PRODUCT_PREFIX):]
    if suffix and not suffix[0].isalnum():
        raise ValueError(
            "Output filename cannot insert '_' or other separators immediately after "
            f"'{ERP_PRODUCT_PREFIX}'. Use letters/numbers directly after the prefix, for example "
            f"'{ERP_PRODUCT_PREFIX}AI0322.xlsx'. Received: {path.name}"
        )


def parse_product_page(url: str) -> Dict[str, str]:
    page = fetch_html(url)

    title = clean_text(
        extract_first(page, [r"<title>(.*?) \| Fruugo UK</title>"], flags=re.S | re.I)
    )
    description = clean_text(
        extract_first(
            page,
            [
                r'"description": "(.*?)",\s*"sku":',
                r'<meta\s+name="description"\s+content="(.*?)"'
            ],
            flags=re.S | re.I
        )
    )
    price = extract_first(page, [r'"price": "(.*?)"'], flags=re.I)
    rrp = extract_first(
        page,
        [
            r"RRP[^£]*£\s*([0-9]+(?:\.[0-9]+)?)",
            r"MSRP[^£]*£\s*([0-9]+(?:\.[0-9]+)?)",
            r'\"rrp\"\s*:\s*\"?([0-9]+(?:\.[0-9]+)?)'
        ],
        flags=re.I
    )
    barcode = extract_first(
        page,
        [
            r'"gtin": "(.*?)"',
            r"<dt[^>]*>\s*EAN\s*</dt>\s*<dd[^>]*>(.*?)</dd>"
        ],
        flags=re.S | re.I
    )
    color = clean_text(
        extract_first(
            page,
            [
                r"<dt[^>]*>\s*Colour\s*</dt>\s*<dd[^>]*>(.*?)</dd>",
                r"<dt[^>]*>\s*Color\s*</dt>\s*<dd[^>]*>(.*?)</dd>"
            ],
            flags=re.S | re.I
        )
    )
    size = clean_text(
        extract_first(page, [r"<dt[^>]*>\s*Size\s*</dt>\s*<dd[^>]*>(.*?)</dd>"], flags=re.S | re.I)
    )
    images = list(dict.fromkeys(re.findall(r"https://img\.fruugo\.com/product/[^\"']+?_max\.jpg", page)))

    category = extract_breadcrumb_category(page)
    if not category:
        breadcrumb_names = [
            html.unescape(name)
            for name in re.findall(r'"name":"(.*?)"', extract_first(
                page,
                [r'"itemListElement":\[(.*?)\]\s*}\s*]\s*}'],
                flags=re.S | re.I
            ))
            if name != "Fruugo"
        ]
        category = normalize_category(" > ".join(breadcrumb_names))
    if not rrp:
        rrp = price

    return {
        "title": title,
        "description": description,
        "rrp": rrp,
        "category": category,
        "color": color,
        "size": size,
        "barcode": barcode,
        "images": "|".join(images[: len(IMAGE_COLUMNS)]),
    }


def normalize_name(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.strip().lower())
    return slug.strip("_") or "product"


def update_query_param(url: str, name: str, value: str) -> str:
    parsed = urlparse(url)
    query_items = [(key, item) for key, item in parse_qsl(parsed.query, keep_blank_values=True) if key != name]
    query_items.append((name, value))
    return urlunparse(parsed._replace(query=urlencode(query_items)))


def extract_sort_value(page: str) -> Tuple[str, str]:
    for select_match in re.finditer(
        r"<select[^>]*name=[\"'](?P<name>[^\"']+)[\"'][^>]*>(?P<body>.*?)</select>",
        page,
        re.S | re.I,
    ):
        field_name = select_match.group("name").strip()
        body = select_match.group("body")
        for option_match in re.finditer(
            r"<option[^>]*value=[\"'](?P<value>[^\"']*)[\"'][^>]*>(?P<label>.*?)</option>",
            body,
            re.S | re.I,
        ):
            label = clean_text(option_match.group("label"))
            if "best selling" in label.lower():
                return field_name, html.unescape(option_match.group("value").strip())
    return "", ""


def parse_sitemap_categories(page: str) -> List[CategoryTarget]:
    categories: List[CategoryTarget] = []
    seen = set()
    for match in re.finditer(
        r"<a[^>]+href=[\"'](?P<href>/[^\"']+/a-\d+(?:[^\"']*)?)[\"'][^>]*>(?P<label>.*?)</a>",
        page,
        re.S | re.I,
    ):
        label = clean_text(match.group("label"))
        if not label:
            continue
        url = urljoin(FRUUGO_BASE_URL, html.unescape(match.group("href")))
        if url in seen:
            continue
        seen.add(url)
        categories.append(CategoryTarget(name=label, url=url))
    return categories


def resolve_category_targets(category_names: List[str]) -> List[CategoryTarget]:
    targets: List[CategoryTarget] = []
    mapped = {normalize_name(name): url for name, url in DEFAULT_HOT_CATEGORY_URLS.items()}
    missing: List[str] = []
    for raw_name in category_names:
        name = raw_name.strip()
        if not name:
            continue
        normalized = normalize_name(name)

        if name.startswith("http://") or name.startswith("https://"):
            targets.append(CategoryTarget(name=name, url=name))
            continue

        exact_url = mapped.get(normalized)
        if exact_url:
            targets.append(CategoryTarget(name=name, url=exact_url))
            continue

        if "=" in name:
            label, raw_url = [part.strip() for part in name.split("=", 1)]
            if raw_url.startswith("http://") or raw_url.startswith("https://"):
                targets.append(CategoryTarget(name=label or raw_url, url=raw_url))
                continue

        missing.append(name)

    if missing:
        raise ValueError(
            "Unknown categories. Use built-in names, full category URLs, or 'Name=https://...' pairs: "
            + ", ".join(missing)
        )
    return targets


def parse_category_product_urls(page: str) -> List[str]:
    urls: List[str] = []
    seen = set()
    for raw_url in re.findall(r"href=[\"'](?P<url>/[^\"']+/p-\d+(?:-\d+)?(?:\?[^\"']*)?)[\"']", page, re.I):
        url = urljoin(FRUUGO_BASE_URL, html.unescape(raw_url))
        if url in seen:
            continue
        seen.add(url)
        urls.append(update_query_param(url, "language", "en"))
    return urls


def parse_category_product_cards(page: str) -> List[Tuple[str, str]]:
    cards: List[Tuple[str, str]] = []
    seen = set()
    for match in re.finditer(
        r"<div[^>]*class=[\"'][^\"']*product-item[^\"']*[\"'][^>]*data-name=[\"'](?P<title>[^\"']+)[\"'][^>]*>"
        r".*?<a\s+href=[\"'](?P<url>/[^\"']+/p-\d+(?:-\d+)?(?:\?[^\"']*)?)[\"']",
        page,
        re.S | re.I,
    ):
        title = clean_text(match.group("title"))
        url = update_query_param(urljoin(FRUUGO_BASE_URL, html.unescape(match.group("url"))), "language", "en")
        if url in seen:
            continue
        seen.add(url)
        cards.append((title, url))
    return cards


def is_relevant_product_title(category_name: str, title: str) -> bool:
    lowered = title.lower()
    keywords = CATEGORY_KEYWORD_HINTS.get(category_name, [])
    return any(keyword in lowered for keyword in keywords)


def collect_category_candidates(target: CategoryTarget, per_category: int) -> List[ProductCandidate]:
    page = fetch_html(target.url)
    sort_name, sort_value = extract_sort_value(page)
    if sort_name and sort_value:
        page = fetch_html(update_query_param(target.url, sort_name, sort_value))

    cards = parse_category_product_cards(page)
    if cards:
        filtered_cards = [card for card in cards if is_relevant_product_title(target.name, card[0])]
        selected_cards = filtered_cards or cards
    else:
        selected_cards = [("", url) for url in parse_category_product_urls(page)]

    return [
        ProductCandidate(category_name=target.name, url=url, rank=index, title=title)
        for index, (title, url) in enumerate(selected_cards[:per_category], start=1)
    ]


def build_discovered_product_input(candidate: ProductCandidate, scraped: Dict[str, str], sequence: int) -> ProductInput:
    return ProductInput(
        group_id=f"hot_{sequence:02d}_{slugify(candidate.category_name)}",
        url=candidate.url,
        is_primary="1",
        color=scraped.get("color", ""),
        size=scraped.get("size", ""),
        barcode=scraped.get("barcode", ""),
        rrp=scraped.get("rrp", ""),
        category=scraped.get("category", "") or candidate.category_name,
        title=scraped.get("title", ""),
        description=scraped.get("description", ""),
        images=scraped.get("images", ""),
    )


def build_default_output_path(args: argparse.Namespace) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    target_dir = Path(args.inventory_dir) if args.inventory_dir else Path.cwd()
    return target_dir / f"{ERP_PRODUCT_PREFIX}AI{timestamp}.xlsx"


def build_default_discovery_csv_path(args: argparse.Namespace) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d")
    if args.output:
        target_dir = Path(args.output).parent
    elif args.inventory_dir:
        target_dir = Path(args.inventory_dir)
    else:
        target_dir = Path.cwd()
    return target_dir / f"fruugo_hot_products_{timestamp}.csv"


def write_discovery_csv(path: Path, rows: List[ProductInput]) -> None:
    fieldnames = ["url", "title", "description", "rrp", "category", "color", "size", "barcode", "images"]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "url": row.url,
                    "title": row.title,
                    "description": row.description,
                    "rrp": row.rrp,
                    "category": row.category,
                    "color": row.color,
                    "size": row.size,
                    "barcode": row.barcode,
                    "images": row.images,
                }
            )


def discover_hot_products(args: argparse.Namespace) -> List[ProductInput]:
    if args.discover_hot_count <= 0:
        raise ValueError("--discover-hot-count must be greater than 0 when --input is omitted")
    if args.discover_products_per_category <= 0:
        raise ValueError("--discover-products-per-category must be greater than 0")

    category_names = [part.strip() for part in args.discover_categories.split(",") if part.strip()]
    if not category_names:
        raise ValueError("No categories were provided for discovery")

    targets = resolve_category_targets(category_names)
    candidate_buckets: List[List[ProductCandidate]] = []
    for index, target in enumerate(targets, start=1):
        print(f"[discover {index}/{len(targets)}] Category {target.name}")
        candidates = collect_category_candidates(target, args.discover_products_per_category)
        if candidates:
            candidate_buckets.append(candidates)

    discovered_rows: List[ProductInput] = []
    seen_urls = set()
    round_index = 0
    while len(discovered_rows) < args.discover_hot_count:
        progressed = False
        for bucket in candidate_buckets:
            if round_index >= len(bucket):
                continue
            candidate = bucket[round_index]
            if candidate.url in seen_urls:
                continue
            sequence = len(discovered_rows) + 1
            print(
                f"[hot {sequence}/{args.discover_hot_count}] "
                f"Scraping {candidate.url} ({candidate.category_name} #{candidate.rank})"
            )
            scraped = parse_product_page(candidate.url)
            discovered_rows.append(build_discovered_product_input(candidate, scraped, sequence))
            seen_urls.add(candidate.url)
            progressed = True
            if len(discovered_rows) >= args.discover_hot_count:
                break
        if not progressed:
            break
        round_index += 1

    if len(discovered_rows) < args.discover_hot_count:
        raise ValueError(
            f"Only discovered {len(discovered_rows)} products, fewer than requested {args.discover_hot_count}"
        )
    return discovered_rows


def merge_product(source: ProductInput, scraped: Dict[str, str]) -> ProductData:
    title = source.title or scraped.get("title", "")
    description = source.description or scraped.get("description", "")
    category = normalize_category(source.category or scraped.get("category", ""))
    color = title_case_words(source.color or scraped.get("color", ""))
    size = source.size or scraped.get("size", "")
    attrs = derive_attrs(color, size, source.attrs)
    barcode = source.barcode or scraped.get("barcode", "") or random_barcode()
    image_string = source.images or scraped.get("images", "")
    images = [part.strip() for part in image_string.split("|") if part.strip()]
    rrp = parse_decimal(source.rrp or scraped.get("rrp", "0"))

    if not title:
        raise ValueError(f"Could not determine title for {source.url}")
    if not description:
        raise ValueError(f"Could not determine description for {source.url}")
    if not category:
        raise ValueError(f"Could not determine category for {source.url}")
    if not images:
        raise ValueError(f"Could not determine images for {source.url}")

    return ProductData(
        title=title,
        description=description,
        rrp=rrp,
        category=category,
        color=color,
        size=size,
        attrs=attrs,
        barcode=barcode,
        images=images[: len(IMAGE_COLUMNS)],
        source_url=source.url
    )


def group_products(source_rows: List[ProductInput]) -> List[List[ProductInput]]:
    grouped: Dict[str, List[ProductInput]] = defaultdict(list)
    order: List[str] = []

    for row in source_rows:
        if row.group_id not in grouped:
            order.append(row.group_id)
        grouped[row.group_id].append(row)

    return [grouped[group_id] for group_id in order]


def finalize_group(rows: List[ProductData], source_rows: List[ProductInput]) -> GroupedProduct:
    primary_index = 0
    for index, source in enumerate(source_rows):
        if parse_bool(source.is_primary):
            primary_index = index
            break

    primary = rows[primary_index]
    group_attrs = set()
    for row in rows:
        if row.color:
            group_attrs.add("COLOR")
        if row.size:
            group_attrs.add("SIZE")
        for attr in row.attrs.split(","):
            cleaned = attr.strip().upper()
            if cleaned:
                group_attrs.add(cleaned)

    ordered_attrs = [name for name in ["COLOR", "SIZE"] if name in group_attrs]
    ordered_attrs.extend(sorted(group_attrs - set(ordered_attrs)))
    merged_attrs = ",".join(ordered_attrs)

    finalized_rows: List[ProductData] = []
    for index, row in enumerate(rows):
        attrs = row.attrs or merged_attrs
        if merged_attrs:
            attrs = merged_attrs
        finalized_rows.append(
            ProductData(
                title=row.title or primary.title,
                description=row.description or primary.description,
                rrp=row.rrp or primary.rrp,
                category=row.category or primary.category,
                color=row.color,
                size=row.size,
                attrs=attrs,
                barcode=row.barcode,
                images=row.images or primary.images,
                source_url=row.source_url
            )
        )

    if primary_index != 0:
        finalized_rows[0], finalized_rows[primary_index] = finalized_rows[primary_index], finalized_rows[0]

    return GroupedProduct(group_id=source_rows[0].group_id, rows=finalized_rows)


def build_vendor_sku(operator: str, shop: str, date_code: str, index: int) -> str:
    timestamp = getattr(build_vendor_sku, "_run_timestamp", "")
    if not timestamp:
        timestamp = datetime.now().strftime("%H%M%S")
        setattr(build_vendor_sku, "_run_timestamp", timestamp)
    return f"{operator}-{shop}-{date_code}{timestamp}-{index:03d}"


def clear_rows(worksheet) -> None:
    for row in range(DATA_START_ROW, worksheet.max_row + 1):
        for column in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row, column=column).value = None


def write_sheet(args: argparse.Namespace, groups: List[GroupedProduct]) -> None:
    workbook = load_workbook(args.template)
    worksheet = workbook.active

    if args.clear_existing:
        clear_rows(worksheet)

    discount_rate = Decimal(args.discount_rate)

    offset = 0
    for group in groups:
        primary_barcode = group.rows[0].barcode
        for index_in_group, product in enumerate(group.rows):
            offset += 1
            row = DATA_START_ROW + offset - 1
            vendor_sku = build_vendor_sku(args.operator, args.shop, args.date_code, offset)
            discount_price = (product.rrp * discount_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            values = {
                "C": product.barcode,
                "D": product.barcode,
                "F": args.vendor_code,
                "G": vendor_sku,
                "H": 0.01,
                "K": product.title,
                "L": product.title,
                "M": product.description,
                "N": 1,
                "O": 1,
                "P": 1,
                "Q": 50,
                "T": "Everything Else",
                "V": 100000,
                "X": None if index_in_group == 0 else primary_barcode,
                "AI": vendor_sku,
                "AJ": product.title,
                "AK": product.description,
                "AL": float(product.rrp),
                "AM": 0.01,
                "AP": "STD",
                "AQ": 0,
                "AR": product.category,
                "AS": product.size,
                "AT": product.color,
                "BH": args.manufacturer,
                "BI": args.brand,
                "BK": 2,
                "BM": float(discount_price),
                "BY": product.attrs,
            }

            for column in EMPTY_COLUMNS:
                values[column] = None

            for column, image_url in zip(IMAGE_COLUMNS, product.images):
                values[column] = image_url

            for column, value in values.items():
                worksheet[f"{column}{row}"] = value

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_inventory_path(args: argparse.Namespace) -> Path:
    target_dir = Path(args.inventory_dir) if args.inventory_dir else Path(args.output).parent
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return target_dir / f"{args.inventory_prefix}{timestamp}.xlsx"


def write_inventory_sheet(args: argparse.Namespace, groups: List[GroupedProduct]) -> Path:
    if args.inventory_min_qty > args.inventory_max_qty:
        raise ValueError("inventory-min-qty cannot be greater than inventory-max-qty")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Worksheet"
    worksheet["A1"] = "code"
    worksheet["B1"] = "codeFieldName"
    worksheet["C1"] = "warehouse"
    worksheet["D1"] = "quantity"

    row = 2
    for group in groups:
        for product in group.rows:
            worksheet[f"A{row}"] = product.barcode
            worksheet[f"B{row}"] = args.inventory_code_field_name
            worksheet[f"C{row}"] = args.inventory_warehouse
            worksheet[f"D{row}"] = random.randint(args.inventory_min_qty, args.inventory_max_qty)
            row += 1

    inventory_path = build_inventory_path(args)
    inventory_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(inventory_path)
    return inventory_path


def main() -> int:
    args = parse_args()

    if not args.input and args.discover_hot_count <= 0:
        raise ValueError("Provide --input or set --discover-hot-count to automatically find hot products")

    source_rows: List[ProductInput]
    discovery_csv_path: Path | None = None
    if args.input:
        input_path = Path(args.input)
        if not input_path.exists():
            raise FileNotFoundError(f"Input CSV not found: {input_path}")
        source_rows = read_input_csv(input_path)
    else:
        source_rows = discover_hot_products(args)
        discovery_csv_path = Path(args.discover_output_csv) if args.discover_output_csv else build_default_discovery_csv_path(args)
        write_discovery_csv(discovery_csv_path, source_rows)
        print(f"Discovered CSV: {discovery_csv_path}")

    if args.discover_only:
        return 0

    if not args.template:
        raise ValueError("--template is required unless --discover-only is used")
    template_path = Path(args.template)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    if not args.output:
        args.output = str(build_default_output_path(args))
    output_path = Path(args.output)
    validate_output_filename(output_path)

    scraped_cache: Dict[str, Dict[str, str]] = {}
    grouped_sources = group_products(source_rows)
    tracker: LinkTracker | None = None
    if args.tracker_db:
        tracker = LinkTracker(Path(args.tracker_db))
        tracker_source_label = args.tracker_source_label or args.input or (str(discovery_csv_path) if discovery_csv_path else "")
        tracker.ensure_urls((source.url for source in source_rows), source_label=tracker_source_label)

        filtered_groups: List[List[ProductInput]] = []
        skipped_count = 0
        for group_sources in grouped_sources:
            url = group_sources[0].url
            if tracker.claim_url(url, allow_failed=not args.tracker_skip_failed):
                filtered_groups.append(group_sources)
            else:
                skipped_count += 1
                print(f"Skip consumed URL: {url}")

        grouped_sources = filtered_groups
        print(f"Tracker claimed groups: {len(grouped_sources)}, skipped consumed groups: {skipped_count}")

    if not grouped_sources:
        print("No unconsumed URLs to process.")
        return 0

    groups: List[GroupedProduct] = []

    for group_index, group_sources in enumerate(grouped_sources, start=1):
        url = group_sources[0].url
        try:
            if url not in scraped_cache and not all(source_has_required_fields(source) for source in group_sources):
                print(f"[{group_index}/{len(grouped_sources)}] Scraping {url}")
                scraped_cache[url] = parse_product_page(url)
            scraped = scraped_cache.get(url, {})
            rows = [merge_product(source, scraped) for source in group_sources]
            groups.append(finalize_group(rows, group_sources))
            if tracker:
                tracker.mark_done([url])
        except Exception as exc:
            if tracker:
                tracker.mark_failed([url], str(exc))
            raise

    write_sheet(args, groups)
    inventory_path = write_inventory_sheet(args, groups)

    print(f"Generated: {args.output}")
    print(f"Inventory: {inventory_path}")
    sku_index = 0
    for group in groups:
        for product in group.rows:
            sku_index += 1
            sku = build_vendor_sku(args.operator, args.shop, args.date_code, sku_index)
            print(f"{sku} <- {product.source_url}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        raise SystemExit(130)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
